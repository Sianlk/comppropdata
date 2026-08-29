from __future__ import annotations

import csv
import io
import json
import os
import re
import xml.etree.ElementTree as ET
from datetime import date, timedelta
from typing import Any
from urllib.parse import quote

import httpx
from fastapi import Depends, HTTPException, Request
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
from sqlalchemy.orm import Session

from . import full_stack as stack
from . import main as core

app = stack.app
ENV = os.getenv("ENV", "development").lower()
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-5.6-terra")
GOOGLE_TRENDS_GEO = os.getenv("GOOGLE_TRENDS_GEO", "GB")
GOOGLE_TRENDS_RSS_URL = os.getenv("GOOGLE_TRENDS_RSS_URL", "https://trends.google.com/trending/rss")
GOOGLE_SEARCH_CONSOLE_ACCESS_TOKEN = os.getenv("GOOGLE_SEARCH_CONSOLE_ACCESS_TOKEN", "")
GOOGLE_SEARCH_CONSOLE_SITE_URL = os.getenv("GOOGLE_SEARCH_CONSOLE_SITE_URL", "")
SEMRUSH_API_KEY = os.getenv("SEMRUSH_API_KEY", "")
SEMRUSH_API_URL = os.getenv("SEMRUSH_API_URL", "https://api.semrush.com/")
PUBLIC_API_URL = os.getenv("PUBLIC_API_URL", "http://localhost:8080").rstrip("/")
SITE_URL = os.getenv("NEXT_PUBLIC_SITE_URL", os.getenv("FRONTEND_URL", "http://localhost:3000")).rstrip("/")
BREVO_API_KEY = os.getenv("BREVO_API_KEY", "")
FROM_EMAIL = os.getenv("FROM_EMAIL", "")
FROM_NAME = os.getenv("FROM_NAME", "Property Development, Augmented")
NOTIFY_EMAIL = os.getenv("NOTIFY_EMAIL", "")

if ENV == "production" and (not stack.JWT_SECRET or stack.JWT_SECRET == "dev-change-me" or "replace-" in stack.JWT_SECRET.lower() or len(stack.JWT_SECRET) < 32):
    raise RuntimeError("Production JWT_SECRET must be a strong private secret (32+ characters).")

CORE_POLICY = """You are part of Property Development, Augmented, a UK property-development and construction decision-support platform.
Use supplied and retrieved evidence carefully. Separate CONFIRMED FACT, ASSUMPTION, PROFESSIONAL OPINION, PROPOSAL/INFERENCE and DECISION.
Never manufacture a source, measurement, approval, credential, valuation, professional opinion, keyword metric or project fact.
If evidence is absent or unclear, state NOT ESTABLISHED. Preserve material qualifications and contradictions.
AI augments the developer/project team; it does not replace planners, architects, engineers, surveyors, solicitors, quantity surveyors, building-control professionals, statutory decision-makers or other qualified advisers.
For live policy, technical, legal, safety, finance and material commercial decisions, tell the user what must be verified and by whom.
"""


def _output_text(data: dict[str, Any]) -> str:
    direct = data.get("output_text")
    if isinstance(direct, str) and direct.strip():
        return direct.strip()
    parts: list[str] = []
    for item in data.get("output", []) or []:
        if not isinstance(item, dict):
            continue
        for content in item.get("content", []) or []:
            if isinstance(content, dict) and content.get("type") in {"output_text", "text"}:
                text = content.get("text")
                if isinstance(text, str) and text.strip():
                    parts.append(text.strip())
    return "\n".join(parts).strip()


async def ai_v2(mode: str, prompt: str, context: Any = None):
    if not core.OPENAI_API_KEY:
        return {"configured": False, "content": "AI provider not configured; set OPENAI_API_KEY on the independent backend."}
    instruction = CORE_POLICY + "\nSPECIALIST MODE\n" + core.MODES.get(mode, core.MODES["report-writer"])
    material = json.dumps(context or {}, default=str, ensure_ascii=False)
    payload = {
        "model": OPENAI_MODEL,
        "instructions": instruction,
        "input": f"SOURCE / PROJECT CONTEXT\n{material[:180000]}\n\nTASK\n{prompt}",
    }
    async with httpx.AsyncClient(timeout=120) as client:
        response = await client.post(
            "https://api.openai.com/v1/responses",
            headers={"Authorization": f"Bearer {core.OPENAI_API_KEY}", "Content-Type": "application/json"},
            json=payload,
        )
    if response.status_code >= 400:
        raise HTTPException(502, f"AI provider error {response.status_code}: {response.text[:500]}")
    data = response.json()
    return {
        "configured": True,
        "content": _output_text(data) or "The AI provider returned no text output.",
        "model": data.get("model", OPENAI_MODEL),
        "usage": data.get("usage"),
        "response_id": data.get("id"),
        "provider": "OpenAI Responses API",
    }


# All existing core routes resolve core.ai at runtime, so this upgrades the provider centrally.
core.ai = ai_v2
core.OPENAI_MODEL = OPENAI_MODEL


@app.middleware("http")
async def project_integrity_guard(request: Request, call_next):
    """Prevent a direct API caller from labelling a variation Approved without explicit approval evidence."""
    if request.method == "POST" and re.fullmatch(r"/api/v1/projects/\d+/registers", request.url.path):
        try:
            body = await request.body()
            payload = json.loads(body or b"{}")
            if str(payload.get("kind", "")).lower() == "variation" and str(payload.get("status", "")).lower() == "approved":
                data = payload.get("data") or {}
                evidence = data.get("approval_evidence") or data.get("approval") or (data.get("approved_by") and data.get("approved_at"))
                if not evidence:
                    return JSONResponse({"detail": "Variation cannot be marked approved without explicit approval evidence."}, status_code=422)
        except json.JSONDecodeError:
            pass
    return await call_next(request)


class TrendsRequest(BaseModel):
    geo: str = "GB"
    limit: int = Field(25, ge=1, le=100)


class SearchConsoleRequest(BaseModel):
    site_url: str = ""
    start_date: str = ""
    end_date: str = ""
    dimensions: list[str] = Field(default_factory=lambda: ["query", "page"])
    row_limit: int = Field(250, ge=1, le=25000)


class KeywordMetricRequest(BaseModel):
    keywords: list[str]
    database: str = "uk"


class LeadCapture(BaseModel):
    email: str
    name: str = ""
    source: str = "site-triage"
    requested_resource: str = "site-triage"
    marketing_consent: bool = False
    payload: dict[str, Any] = Field(default_factory=dict)


class ConsultancySubmit(BaseModel):
    email: str
    name: str = ""
    company: str = ""
    project: str = ""
    problem: str
    deadline: str = ""
    desired_outcome: str = ""
    documents: str = ""
    data_restrictions: str = ""
    marketing_consent: bool = False


def _valid_email(value: str) -> str:
    value = value.strip().lower()
    if len(value) > 320 or not re.fullmatch(r"[^\s@]+@[^\s@]+\.[^\s@]+", value):
        raise HTTPException(422, "Enter a valid email address")
    return value


async def _brevo_send(to_email: str, subject: str, html: str) -> dict[str, Any]:
    if not BREVO_API_KEY or not FROM_EMAIL:
        return {"configured": False, "sent": False}
    payload = {
        "sender": {"name": FROM_NAME, "email": FROM_EMAIL},
        "to": [{"email": to_email}],
        "subject": subject,
        "htmlContent": html,
    }
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.post("https://api.brevo.com/v3/smtp/email", headers={"api-key": BREVO_API_KEY, "Content-Type": "application/json"}, json=payload)
    return {"configured": True, "sent": response.status_code < 300, "status_code": response.status_code}


def _provider_status() -> dict[str, Any]:
    return {
        "api": True,
        "ai": bool(core.OPENAI_API_KEY),
        "ai_model": OPENAI_MODEL if core.OPENAI_API_KEY else None,
        "database": bool(os.getenv("DATABASE_URL", "")),
        "stripe": bool(os.getenv("STRIPE_SECRET_KEY", "")),
        "transactional_email": bool(BREVO_API_KEY and FROM_EMAIL),
        "companies_house": bool(core.COMPANIES_HOUSE_API_KEY),
        "epc": bool(core.EPC_API_URL),
        "google_trends": True,
        "google_search_console": bool(GOOGLE_SEARCH_CONSOLE_ACCESS_TOKEN and GOOGLE_SEARCH_CONSOLE_SITE_URL),
        "semrush": bool(SEMRUSH_API_KEY),
    }


@app.get("/api/v1/system/status")
def system_status():
    return {
        "status": "healthy",
        "service": core.APP,
        "version": core.VERSION,
        "providers": _provider_status(),
        "principles": [
            "Live data must retain source and retrieval context.",
            "Missing evidence is not established.",
            "AI output is working material until human review.",
            "Professional and statutory judgement remains with the appropriate person or authority.",
        ],
    }


# Stable aliases used by the independent frontend. The older endpoints remain available.
@app.post("/api/v1/ai/analyse")
async def ai_analyse_alias(r: core.Analysis):
    return await core.assist(r)


@app.post("/api/v1/calculators/residual")
def residual_alias(r: core.Residual):
    return core.residual(r)


@app.post("/api/v1/quotes/compare")
async def quote_alias(r: core.Quotes):
    return await core.quotes(r)


@app.post("/api/v1/seo/strategy")
async def seo_alias(r: core.SEO):
    return await core.seo(r)


@app.get("/api/v1/data/companies-house/search")
async def companies_house_search(q: str):
    if not core.COMPANIES_HOUSE_API_KEY:
        raise HTTPException(503, "Companies House is not configured")
    url = "https://api.company-information.service.gov.uk/search/companies"
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.get(url, params={"q": q, "items_per_page": 50}, auth=(core.COMPANIES_HOUSE_API_KEY, ""))
        response.raise_for_status()
        data = response.json()
    return {
        "items": data.get("items", []),
        "source": core.src("Companies House API", url, "Verify company status and filings in the authoritative Companies House record."),
    }


@app.post("/api/v1/seo/trending-now")
async def trending_now(r: TrendsRequest):
    url = GOOGLE_TRENDS_RSS_URL
    try:
        async with httpx.AsyncClient(timeout=30, follow_redirects=True) as client:
            response = await client.get(url, params={"geo": r.geo or GOOGLE_TRENDS_GEO}, headers={"User-Agent": "PropertyDevelopmentAugmented/2.0"})
            response.raise_for_status()
        root = ET.fromstring(response.content)
        items = []
        for item in root.findall(".//item")[: r.limit]:
            row: dict[str, Any] = {
                "title": (item.findtext("title") or "").strip(),
                "published": (item.findtext("pubDate") or "").strip(),
                "link": (item.findtext("link") or "").strip(),
            }
            for child in list(item):
                tag = child.tag.split("}")[-1]
                if tag in {"approx_traffic", "picture", "picture_source"}:
                    row[tag] = (child.text or "").strip()
            if row["title"]:
                items.append(row)
        return {
            "geo": r.geo,
            "items": items,
            "source": core.src("Google Trends - Trending now", f"https://trends.google.com/trending?geo={quote(r.geo)}", "Trending searches are not keyword demand forecasts and should not be treated as SEO search volume."),
        }
    except Exception as exc:
        raise HTTPException(502, f"Google Trends feed unavailable: {exc}") from exc


@app.post("/api/v1/seo/search-console")
async def search_console(r: SearchConsoleRequest):
    token = GOOGLE_SEARCH_CONSOLE_ACCESS_TOKEN
    site_url = r.site_url or GOOGLE_SEARCH_CONSOLE_SITE_URL
    if not token or not site_url:
        raise HTTPException(503, "Google Search Console is not configured")
    end = r.end_date or date.today().isoformat()
    start = r.start_date or (date.today() - timedelta(days=28)).isoformat()
    url = f"https://www.googleapis.com/webmasters/v3/sites/{quote(site_url, safe='')}/searchAnalytics/query"
    payload = {"startDate": start, "endDate": end, "dimensions": r.dimensions, "rowLimit": r.row_limit}
    async with httpx.AsyncClient(timeout=45) as client:
        response = await client.post(url, headers={"Authorization": f"Bearer {token}"}, json=payload)
        if response.status_code >= 400:
            raise HTTPException(502, f"Search Console API error {response.status_code}: {response.text[:500]}")
        data = response.json()
    return {
        "site_url": site_url,
        "date_range": {"start": start, "end": end},
        "rows": data.get("rows", []),
        "response_aggregation_type": data.get("responseAggregationType"),
        "source": core.src("Google Search Console Search Analytics", "https://developers.google.com/webmaster-tools/v1/searchanalytics/query", "First-party search performance for an authorised property; data is not a promise of future ranking."),
    }


async def _semrush_phrase(keyword: str, database: str) -> dict[str, Any]:
    if not SEMRUSH_API_KEY:
        return {"keyword": keyword, "configured": False, "volume": None, "cpc": None, "competition": None, "results": None, "trend": None}
    params = {
        "type": "phrase_this",
        "key": SEMRUSH_API_KEY,
        "phrase": keyword,
        "database": database,
        "export_columns": "Ph,Nq,Cp,Co,Nr,Td",
        "display_limit": 1,
    }
    async with httpx.AsyncClient(timeout=30) as client:
        response = await client.get(SEMRUSH_API_URL, params=params)
        if response.status_code >= 400:
            return {"keyword": keyword, "configured": True, "error": f"HTTP {response.status_code}"}
        text = response.text.strip()
    if not text or text.startswith("ERROR"):
        return {"keyword": keyword, "configured": True, "error": text or "No data"}
    rows = list(csv.DictReader(io.StringIO(text), delimiter=";"))
    if not rows:
        return {"keyword": keyword, "configured": True, "volume": None}
    row = rows[0]
    return {
        "keyword": keyword,
        "configured": True,
        "volume": row.get("Search Volume") or row.get("Nq"),
        "cpc": row.get("CPC") or row.get("Cp"),
        "competition": row.get("Competition") or row.get("Co"),
        "results": row.get("Number of Results") or row.get("Nr"),
        "trend": row.get("Trends") or row.get("Td"),
    }


@app.post("/api/v1/seo/measured-keywords")
async def measured_keywords(r: KeywordMetricRequest):
    metrics = []
    for keyword in list(dict.fromkeys(k.strip() for k in r.keywords if k.strip()))[:50]:
        metrics.append(await _semrush_phrase(keyword, r.database))
    return {
        "database": r.database,
        "metrics": metrics,
        "provider": "Semrush" if SEMRUSH_API_KEY else "not configured",
        "note": "Only provider-returned metrics are surfaced. No volume or competition values are fabricated.",
    }


@app.get("/api/v1/data/epc/status")
def epc_status():
    return {
        "configured": bool(core.EPC_API_URL),
        "service": "Get energy performance of buildings data",
        "source_url": "https://get-energy-performance-data.communities.gov.uk/",
        "note": "The current England and Wales service requires authorised API access. The retired legacy endpoint is not used.",
    }


def _site_triage_pdf() -> bytes:
    questions = [
        "What exactly is the site and ownership extent?",
        "What is the current lawful use?",
        "What is the intended development outcome?",
        "What does the planning history actually establish?",
        "Which current planning policies are obviously engaged?",
        "Are there heritage, flood, ecology, tree, access or other obvious constraints?",
        "What is assumed about access, servicing, parking, cycles and refuse?",
        "What is assumed about neighbours and amenity?",
        "What do we actually know about the existing building?",
        "Which dimensions, levels or site facts are still unverified?",
        "Which cost assumptions are not evidenced?",
        "Which technical reports might be required — and what question would each answer?",
        "Which professional opinions are needed before further commitment?",
        "What is the single biggest unknown?",
        "What is the next cheapest action that removes the most important uncertainty?",
    ]
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("title", parent=styles["Title"], fontSize=24, leading=28, textColor=colors.HexColor("#222222"), spaceAfter=10)
    sub = ParagraphStyle("sub", parent=styles["BodyText"], fontSize=11, leading=16, textColor=colors.HexColor("#6B6862"), spaceAfter=10)
    qstyle = ParagraphStyle("q", parent=styles["BodyText"], fontSize=10, leading=14, spaceAfter=7)
    story = [Paragraph("THE 30-MINUTE AI SITE TRIAGE", title), Paragraph("15 questions to ask before spending thousands investigating a development opportunity", sub), Paragraph("SITE → BUILD → PROVE", sub), Spacer(1, 4 * mm), Paragraph("Use four labels for every material point: <b>CONFIRMED</b> · <b>ASSUMED</b> · <b>UNKNOWN</b> · <b>MATERIAL IF WRONG</b>.", qstyle)]
    for i, question in enumerate(questions, 1):
        story.append(Paragraph(f"<b>{i}. {question}</b><br/>Status: ____________________ &nbsp;&nbsp; Evidence/source: ____________________<br/>Next verification action: __________________________________________________", qstyle))
    story.extend([Spacer(1, 5 * mm), Paragraph("AI should organise uncertainty, not manufacture certainty. If something is not supported by the material, record <b>NOT ESTABLISHED</b>.", qstyle), Paragraph("Educational decision-support material only. Verify title/legal, planning, technical, cost, valuation, tax, finance and safety matters with the appropriate authoritative source and qualified professional.", sub)])
    doc.build(story)
    return buf.getvalue()


@app.get("/api/v1/resources/site-triage.pdf")
def site_triage_resource():
    raw = _site_triage_pdf()
    return StreamingResponse(io.BytesIO(raw), media_type="application/pdf", headers={"Content-Disposition": 'attachment; filename="30-minute-ai-site-triage.pdf"'})


@app.post("/api/v1/leads/capture")
async def lead_capture(r: LeadCapture, s: Session = Depends(stack.db)):
    email = _valid_email(r.email)
    lead = stack.Lead(email=email, name=r.name.strip(), source=r.source, consent=r.marketing_consent, payload_json=json.dumps({**r.payload, "requested_resource": r.requested_resource}))
    s.add(lead)
    s.commit()
    s.refresh(lead)
    download_path = "/api/v1/resources/site-triage.pdf" if r.requested_resource == "site-triage" else None
    delivery = {"configured": False, "sent": False}
    if download_path:
        delivery = await _brevo_send(email, "Your 30-Minute AI Site Triage", f"<p>Thanks {r.name or ''}. Your Site Triage is ready.</p><p><a href=\"{PUBLIC_API_URL}{download_path}\">Download the 30-Minute AI Site Triage</a></p><p>AI should organise uncertainty, not replace professional judgement.</p>")
    return {"id": lead.id, "status": "captured", "marketing_consent": r.marketing_consent, "download_url": download_path, "email_delivery": delivery}


@app.post("/api/v1/consultancy/submit")
async def consultancy_submit(r: ConsultancySubmit, s: Session = Depends(stack.db)):
    email = _valid_email(r.email)
    payload = r.model_dump(exclude={"email", "name", "marketing_consent"})
    lead = stack.Lead(email=email, name=r.name.strip(), source="consultancy-intake", consent=r.marketing_consent, payload_json=json.dumps(payload))
    s.add(lead)
    s.commit()
    s.refresh(lead)
    notice = {"configured": False, "sent": False}
    if NOTIFY_EMAIL:
        notice = await _brevo_send(NOTIFY_EMAIL, f"New consultancy enquiry — {r.name or email}", f"<h3>New Property Development, Augmented enquiry</h3><p><b>Name:</b> {r.name}</p><p><b>Email:</b> {email}</p><p><b>Company:</b> {r.company}</p><p><b>Project:</b> {r.project}</p><p><b>Problem:</b> {r.problem}</p><p><b>Deadline:</b> {r.deadline}</p><p><b>Desired outcome:</b> {r.desired_outcome}</p>")
    return {"id": lead.id, "status": "received", "notification": notice}


def _developer_os_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    headers = ["PROPERTY DEVELOPMENT, AUGMENTED — AI PROPERTY DEVELOPER OS", "SITE → BUILD → PROVE"]
    ws["A1"], ws["A2"] = headers
    ws["A1"].font = Font(bold=True, size=18, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="222222")
    ws.merge_cells("A1:H1"); ws.merge_cells("A2:H2")
    ws["A2"].font = Font(bold=True, color="A4845C")
    ws["A4"] = "Use the registers as verified project records. AI output is working material until reviewed."
    ws.merge_cells("A4:H4")
    definitions = {
        "Site_Triage": ["ID", "Question / issue", "Answer / assumption", "Evidence / source", "Status", "Material if wrong?", "Next action", "Owner / date"],
        "Planning_Evidence": ["ID", "Issue", "Project proposition", "Source", "Date / revision", "Authority / consultee position", "Evidence strength", "Gap / contradiction", "Action", "Status"],
        "Quote_Comparison": ["Work package", "Bidder A", "Bidder B", "Bidder C", "Inclusions / exclusions", "Clarification", "Decision / next action"],
        "Risk_Register": ["ID", "Cause", "Event", "Consequence", "Probability 1-5", "Impact 1-5", "Score", "Mitigation", "Owner", "Status"],
        "Variation_Tracker": ["ID", "Date", "Description", "Reason", "Scope / drawing ref", "Value", "Programme impact", "Approval status", "Approval evidence", "Next action"],
        "Decision_Log": ["ID", "Decision required", "Evidence reviewed", "Decision", "Decision maker", "Date", "Conditions / rationale", "Actions", "Source refs", "Status"],
        "Document_Register": ["ID", "Discipline", "Document type", "Title", "Revision", "Date", "Author / source", "Current?", "Supersedes", "Location / link", "Notes"],
        "Weekly_Report": ["Section", "Current position", "Owner / action"],
        "Prompt_Library": ["Workflow", "Purpose", "Human verification rule", "Suggested specialist mode"],
    }
    for name, cols in definitions.items():
        sh = wb.create_sheet(name)
        sh.append(cols)
        for cell in sh[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="222222")
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for idx, col in enumerate(cols, 1):
            sh.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = min(max(len(col) + 4, 14), 34)
        if name == "Risk_Register":
            for row in range(2, 102):
                sh.cell(row, 7, f'=IF(OR(E{row}="",F{row}=""),"",E{row}*F{row})')
        if name == "Variation_Tracker":
            sh["H2"] = "Approval not evidenced"
        sh.freeze_panes = "A2"
    prompts = wb["Prompt_Library"]
    for row in [
        ["Site opportunity triage", "Separate fact, assumption, unknown and material-if-wrong issues", "Verify primary sources and professional questions", "Site Analyst"],
        ["Planning evidence matrix", "Map each material proposition to evidence", "Current policy and planning judgement require verification", "Planning Evidence Analyst"],
        ["Quote normalisation", "Compare scope rather than headline total", "Preserve bidder wording; do not appoint/certify", "Procurement Analyst"],
        ["Variation control", "Record scope/cost/programme/evidence", "Never infer approval", "Project Controls Analyst"],
        ["Weekly reporting", "Draft narrative from maintained registers", "Do not invent causes or facts", "Report Writer"],
    ]:
        prompts.append(row)
    out = io.BytesIO(); wb.save(out); return out.getvalue()


@app.get("/api/v1/products/ai-property-developer-os/generated-download")
def generated_os(user: stack.User = Depends(stack.me), s: Session = Depends(stack.db)):
    entitled = user.is_admin or s.query(stack.Purchase).filter(stack.Purchase.user_id == user.id, stack.Purchase.product_slug == "ai-property-developer-os", stack.Purchase.status == "paid").first()
    if not entitled:
        raise HTTPException(403, "No paid entitlement for AI Property Developer OS")
    raw = _developer_os_xlsx()
    return StreamingResponse(io.BytesIO(raw), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="AI-Property-Developer-OS.xlsx"'})
